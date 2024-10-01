if __name__ == '__main__':
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    ## Exercise 1**: Create a new Excel workbook and add a worksheet to it using openpyxl.

    wb = Workbook()

    newsheet = wb.create_sheet("openpysheet", 0)

    ## Exercise 2**: Write "Hello, World!" into cell A1 of the worksheet.

    cell = newsheet['A1']
    cell.value = 'Hello World'

    ## Exercise 3**: Change the font style and size of the text in cell A1.

    cell.font = Font(name='Candara', sz=30)

    ## Exercise 4**: Create a new worksheet and add data to it.
    ## For example, add a simple table of names and ages.

    data_sheet = wb.create_sheet('Data', 0)

    treeData = [
        ['Name', 'Goals'], ['Cristiano Ronaldo', 525], ['Lionel Messi', 497],
        ['Pelé', 604], ['Romàrio', 544], ['Ferenc Puskás', 515], ['Josef Bican*', 515],
        ['Jimmy Jones', 332], ['Gerd Müller', 405], ['Joe Bambrick', 348], ['Abe Lenstra', 573]
    ]
    # range = ['B2:C12']
    start_row = 2
    start_col = 2
    for i, row_data in enumerate(treeData):
        for j, value in enumerate(row_data):
            data_sheet.cell(row=start_row + i, column=start_col + j, value=value)

    # **Exercise 5**: Change the background color of the header row in the table
    color = PatternFill(fill_type='solid', fgColor='C42F2F')
    font = Font(name='Candara', bold=True, size=14, color='FFFFFF')
    for row in data_sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.fill = color
            cell.font = font

    # ** Exercise  6 **: Apply borders to the cells in the table to create a border around the data.

    internal_border = Border(left=Side(style='thin', border_style='dashDot'),
                             right=Side(style='thin', border_style='dashDot'),
                             top=Side(style='thin', border_style='dashDot'),
                             bottom=Side(style='thin', border_style='dashDot')
                             )

    for row in data_sheet.iter_rows(min_row=2, max_row=12, min_col=2, max_col=3):
        for cell in row:
            cell.border = internal_border

    # ** Exercise 7 **: Merge cells in a specific range(e.g., merge cells A2 to B2).

    # cell_to_merge = data_sheet['A1:B1']

    data_sheet.merge_cells(range_string='A1:B1')
    data_sheet['A1'].fill = color
if __name__ == '__main__':
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    ## Exercise 1**: Create a new Excel workbook and add a worksheet to it using openpyxl.

    wb = Workbook()

    newsheet = wb.create_sheet("openpysheet", 0)

    ## Exercise 2**: Write "Hello, World!" into cell A1 of the worksheet.

    cell = newsheet['A1']
    cell.value = 'Hello World'

    ## Exercise 3**: Change the font style and size of the text in cell A1.

    cell.font = Font(name='Candara', sz=30)

    ## Exercise 4**: Create a new worksheet and add data to it.
    ## For example, add a simple table of names and ages.

    data_sheet = wb.create_sheet('Data', 0)

    treeData = [
        ['Name', 'Goals'], ['Cristiano Ronaldo', 525], ['Lionel Messi', 497],
        ['Pelé', 604], ['Romàrio', 544], ['Ferenc Puskás', 515], ['Josef Bican*', 515],
        ['Jimmy Jones', 332], ['Gerd Müller', 405], ['Joe Bambrick', 348], ['Abe Lenstra', 573]
    ]
    # range = ['B2:C12']
    start_row = 2
    start_col = 2
    for i, row_data in enumerate(treeData):
        for j, value in enumerate(row_data):
            data_sheet.cell(row=start_row + i, column=start_col + j, value=value)

    # **Exercise 5**: Change the background color of the header row in the table
    color = PatternFill(fill_type='solid', fgColor='C42F2F')
    font = Font(name='Candara', bold=True, size=14, color='FFFFFF')
    for row in data_sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.fill = color
            cell.font = font

    # ** Exercise  6 **: Apply borders to the cells in the table to create a border around the data.

    internal_border = Border(left=Side(style='thin', border_style='dashDot'),
                             right=Side(style='thin', border_style='dashDot'),
                             top=Side(style='thin', border_style='dashDot'),
                             bottom=Side(style='thin', border_style='dashDot')
                             )

    for row in data_sheet.iter_rows(min_row=2, max_row=12, min_col=2, max_col=3):
        for cell in row:
            cell.border = internal_border

    # ** Exercise 7 **: Merge cells in a specific range.

    data_sheet.merge_cells(range_string='A1:B1')
    data_sheet['A1'].fill = color

    # Save
    wb.save(r"C:\Users\karim\Downloads\Exercises.xlsx")
